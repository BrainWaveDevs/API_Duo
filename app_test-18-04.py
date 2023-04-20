#━━━━━━━━━❮Bibliotecas❯━━━━━━━━━
from flask import Flask, jsonify, request, send_file, render_template
from tinydb import TinyDB, Query
from openpyxl import Workbook
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

app = Flask(__name__)

#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Criação do banco de dados
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


db = TinyDB('db.json')
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Criação das tabelas

alunos_table = db.table('alunos')
cursos_table = db.table('cursos')
presenca_table = db.table('presenca')


# Definição da classe Aluno para representar os objetos de aluno
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

#━━━━━━━━━━━━━━❮◆❯━━━━━━━━━━━━━━

class Aluno:
    def __init__(self, nome, idade, curso):
        self.nome = nome
        self.idade = idade
        self.curso = curso
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Definição da classe Curso para representar os objetos de curso
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class Curso:
    def __init__(self, nome, codigo):
        self.nome = nome
        self.codigo = codigo
#━━━━━━━━━━━━━━❮◆❯━━━━━━━━━━━━━━


#━━━━━━━━━━━━━━❮ROTAS❯━━━━━━━━━━━━━━

# Rota para TESTE HELLO WORLD
@app.route('/hello', methods=['GET'])
def HelloWorld():
    return 'Hello World'

# Rota para a página inicial
@app.route('/', methods=['GET', 'POST'])
def home():
    return render_template('index.html')
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para consultar todos os alunos
@app.route('/alunos', methods=['GET'])
def get_alunos():
    alunos = alunos_table.all()
    return jsonify(alunos)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para consultar um aluno por ID (ROTA 1)
@app.route('/alunos/<int:aluno_id>', methods=['GET'])
def get_aluno(aluno_id):
    aluno = alunos_table.get(doc_id=aluno_id)
    if aluno:
        return jsonify(aluno)
    else:
        return jsonify({'error': 'Aluno não encontrado'}), 404
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para cadastrar um novo aluno (ROTA 2)
@app.route('/alunos', methods=['POST'])
def add_aluno():
    data = request.get_json()
    nome = data['nome']
    idade = data['idade']
    curso_id = data['curso_id']
    #http://127.0.0.1:5000
    curso = cursos_table.get(doc_id=curso_id)
    if not curso:
        return jsonify({'error': 'Curso não encontrado'}), 404
    
    aluno = Aluno(nome, idade, curso)
    aluno_id = alunos_table.insert({'nome': aluno.nome, 'idade': aluno.idade, 'curso': aluno.curso.__dict__})
    
    return jsonify({'aluno_id': aluno_id}), 201
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para atualizar um aluno por ID (ROTA 3)
@app.route('/alunos/<int:aluno_id>', methods=['PUT'])
def update_aluno(aluno_id):
    data = request.get_json()
    nome = data['nome']
    idade = data['idade']
    curso_id = data['curso_id']
    
    aluno = Aluno(nome, idade, curso_id)
    alunos_table.update({'nome': aluno.nome, 'idade': aluno.idade, 'curso': aluno.curso.__dict__}, doc_ids=[aluno_id])
    
    return jsonify({'message': 'Aluno atualizado com sucesso'}), 200
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para excluir um aluno por ID (ROTA 4)
@app.route('/alunos/<int:aluno_id>', methods=['DELETE'])
def delete_aluno(aluno_id):
    alunos_table.remove(doc_ids=[aluno_id])
    presenca_table.remove(Query().aluno_id == aluno_id)
    return '', 204
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para consultar todos os cursos (ROTA 5)
@app.route('/cursos', methods=['GET'])
def get_cursos():
    cursos = cursos_table.all()
    return jsonify(cursos)
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para consultar um curso por ID (ROTA 6)
@app.route('/cursos/<int:curso_id>', methods=['GET'])
def get_curso(curso_id):
    curso = cursos_table.get(doc_id=curso_id)
    if curso:
        return jsonify(curso)
    else:
        return jsonify({'error': 'Curso não encontrado'}), 404
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para cadastrar um novo curso (ROTA 7)
@app.route('/cursos', methods=['POST'])
def add_curso():
    data = request.get_json()
    nome = data['nome']
    codigo = data['codigo']
    
    curso = Curso(nome, codigo)
    curso_id = cursos_table.insert({'nome': curso.nome, 'codigo': curso.codigo})
    
    return jsonify({'curso_id': curso_id}), 201
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para atualizar um curso por ID (ROTA 8)
@app.route('/cursos/<int:curso_id>', methods=['PUT'])
def update_curso(curso_id):
    data = request.get_json()
    nome = data['nome']
    codigo = data['codigo']
    
    cursos_table.update({'nome': nome, 'codigo': codigo}, doc_ids=[curso_id])
    
    return jsonify({'message': 'Curso atualizado com sucesso'}), 200
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para excluir um curso por ID (ROTA 9)
@app.route('/cursos/<int:curso_id>', methods=['DELETE'])
def delete_curso(curso_id):
    cursos_table.remove(doc_ids=[curso_id])
    alunos_table.update({'curso': None}, Query().curso.doc_id == curso_id)
    
    return '', 204
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para marcar presença de um aluno em uma aula (ROTA 10)
@app.route('/presenca', methods=['POST'])
def marcar_presenca():
    data = request.get_json()
    aluno_id = data['aluno_id']
    curso_id = data['curso_id']
    
    aluno = alunos_table.get(doc_id=aluno_id)
    if not aluno:
        return jsonify({'error': 'Aluno não encontrado'}), 404
    
    curso = cursos_table.get(doc_id=curso_id)
    if not curso:
        return jsonify({'error': 'Curso não encontrado'}), 404
    
    presenca_table.insert({'aluno_id': aluno_id, 'curso_id': curso_id})
    
    return jsonify({'message': 'Presença registrada com sucesso'}), 201
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Rota para gerar relatório de presença em formato Excel (ROTA 11)
@app.route('/presenca/relatorio', methods=['GET'])
def gerar_relatorio_presenca():
    data = []
    for presenca in presenca_table:
        aluno_id = presenca['aluno_id']
        curso_id = presenca['curso_id']
        aluno = alunos_table.get(doc_id=aluno_id)
        curso = cursos_table.get(doc_id=curso_id)
        if aluno and curso:
            data.append({'aluno_id': aluno_id, 'aluno_nome': aluno['nome'], 'curso_nome': curso['nome']})
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


#━━━━━━━━━━━━━━❮◆❯━━━━━━━━━━━━━━

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Relatório de Presença'
    sheet.cell(row=1, column=1, value='ID Aluno')
    sheet.cell(row=1, column=2, value='Nome Aluno')
    sheet.cell(row=1, column=3, value='Nome Curso')
    for i, presenca in enumerate(data, start=2):
        sheet.cell(row=i, column=1, value=presenca['aluno_id'])
        sheet.cell(row=i, column=2, value=presenca['aluno_nome'])
        sheet.cell(row=i, column=3, value=presenca['curso_nome'])
    
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Definindo o nome do arquivo para o relatório
    filename = 'relatorio_presenca.xlsx'

    # Salvando o arquivo do relatório
    workbook.save(filename)

    # Enviando o arquivo do relatório como resposta para download
    return send_file(filename, as_attachment=True)


#━━━━━━━━━━━━━━❮◆❯━━━━━━━━━━━━━━
if __name__ == '__main__':
     app.run(debug=True)